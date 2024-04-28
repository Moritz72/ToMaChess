import os
import os.path
from typing import Callable, Sequence, cast
from csv import reader
from dbfread import DBF
from openpyxl import load_workbook
from requests import exceptions, get
from zipfile import BadZipFile, ZipFile
from xml.etree.ElementTree import parse
from xlrd import open_workbook
from .player import Player
from ..collection.collection import Collection
from ..common.functions_util import get_app_data_directory, get_uuid_from_numbers, read_file
from ..database.db_collection import DB_COLLECTION, collection_exists
from ..database.db_player import DB_PLAYER


INDICES = [
    (1, 0), (1, 1), (1, 2),
    (2, 0),
    (3, 0), (3, 1), (3, 2),
    (3, 3), (3, 4), (3, 5),
    (4, 0), (4, 1), (4, 2),
    (5, 0), (5, 1),
    (6, 0),
    (7, 0),
    (8, 0), (8, 1), (8, 2),
    (9, 0), (9, 1),
    (10, 0), (10, 1)
]
NAMES_DICT: dict[tuple[int, int], str] = {
    (1, 0): "FIDE (Standard)", (1, 1): "FIDE (Rapid)", (1, 2): "FIDE (Blitz)",
    (2, 0): "DSB",
    (3, 0): "USCF (Regular)", (3, 1): "USCF (Quick)", (3, 2): "USCF (Blitz)",
    (3, 3): "USCF (Online Regular)", (3, 4): "USCF (Online Quick)", (3, 5): "USCF (Online Blitz)",
    (4, 0): "ECF (Standard)", (4, 1): "ECF (Rapid)", (4, 2): "ECF (Blitz)",
    (5, 0): "ÖSB (Standard)", (5, 1): "ÖSB (Rapid)",
    (6, 0): "KNSB",
    (7, 0): "MCF",
    (8, 0): "CFR (Standard)", (8, 1): "CFR (Rapid)", (8, 2): "CFR (Blitz)",
    (9, 0): "FCF (Standard)", (9, 1): "FCF (Blitz)",
    (10, 0): "NZCF (Standard)", (10, 1): "NZCF (Rapid)"
}
FEDERATIONS = [
    "", "Fédération Internationale des Échecs", "Deutscher Schachbund", "United States Chess Federation",
    "English Chess Federation", "Österreichischer Schachbund", "Koninklijke Nederlandse Schaakbond",
    "Persekutuan Catur Malaysia", "Федерация шахмат России", "Suomen Shakkiliitto", "New Zealand Chess Federation"
]
COUNTRIES = [
    "", "", "Germany", "USA", "England", "Austria", "Netherlands", "Malaysia", "Russia", "Finland", "New Zealand"
]


def renew_certificate(certificate: str) -> None:
    if not os.path.exists(os.path.dirname(certificate)):
        os.mkdir(os.path.dirname(certificate))
    url = f"https://raw.githubusercontent.com/Moritz72/ToMaChess/master/certificates/{os.path.basename(certificate)}"
    try:
        response = get(url)
        response.raise_for_status()
        with open(certificate, 'wb') as file:
            file.write(response.content)
    except exceptions.RequestException:
        return


def process_xml(xml_file_path: str, keys: Sequence[str]) -> list[tuple[str, ...]]:
    return [
        tuple(cast(str, child_elem.text) for child_elem in player_elem if child_elem.tag in keys)
        for player_elem in parse(xml_file_path).getroot().iter("player")
    ]


def process_csv(csv_file_path: str, keys: Sequence[str], encoding: str | None = None) -> list[tuple[str, ...]]:
    with open(csv_file_path, 'r', encoding=encoding) as file:
        csv_reader = reader(file)
        header = next(csv_reader)
        indices = tuple(-1 if key == "" else header.index(key) for key in keys)
        return [tuple("" if value == "" else row[indices[j]] for j, value in enumerate(keys)) for row in csv_reader]


def process_xls(xls_file_path: str, keys: Sequence[str]) -> list[tuple[str, ...]]:
    sheet = open_workbook(xls_file_path).sheet_by_index(0)
    header = sheet.row_values(0)
    indices = tuple(-1 if key == "" else header.index(key) for key in keys)
    return [
        tuple("" if value == "" else str(sheet.cell_value(row, indices[j])) for j, value in enumerate(keys))
        for row in range(1, sheet.nrows)
    ]


def process_xlsx(
        xlsx_file_path: str, keys: Sequence[str], header_rows: int = 1, header_add: Sequence[str] | None = None
) -> list[tuple[str, ...]]:
    sheet = load_workbook(xlsx_file_path).active
    header = [cell.value for cell in sheet[1]]
    for row in range(2, header_rows + 1):
        new_header = header.copy()
        values = [cell.value for cell in sheet[row]]
        for j, cell in enumerate(sheet[row]):
            if cell.value is None:
                continue
            left = j + 1 - values[:j + 1][::-1].index(None) if None in values[:j + 1] else 0
            right = j + values[j:].index(None) if None in values[j:] else len(values)
            one_left = [item for item in header[left:right] if item is not None]
            if len(one_left) != 1:
                continue
            new_header[j] = f"{one_left[0]}_{cell.value}"
        header = new_header
    if header_add is not None:
        for i, add in enumerate(header_add):
            if bool(add):
                header[i] = add
    indices = tuple(-1 if key == "" else header.index(key) for key in keys)
    return [
        tuple("" if value == "" else str(row[indices[j]]) for j, value in enumerate(keys))
        for row in sheet.iter_rows(min_row=header_rows + 1, values_only=True)
    ]


def get_link_from_url(search_url: str, search_text: str, front: str = '"', back: str = '"') -> str | None:
    try:
        response = get(search_url, headers={"User-Agent": "XY"})
        response.raise_for_status()
        text = response.text
        if search_text not in text:
            return None
        index = text.index(search_text)
        start = text[:index].rindex(front) + 1
        end = index + text[index:].index(back)
        return text[start:end]
    except exceptions.RequestException:
        return None


def download_file(url: str, file_name: str, verify: bool | str = True, retry: bool = True, agent: bool = True) -> bool:
    download_path = os.path.join(get_app_data_directory(), "temp", file_name)
    try:
        response = get(url, verify=verify, headers={"User-Agent": "XY"} if agent else None)
        response.raise_for_status()
        with open(download_path, 'wb') as file:
            file.write(response.content)
            return True
    except (exceptions.SSLError, OSError):
        if isinstance(verify, str):
            renew_certificate(verify)
        if retry:
            return download_file(url, file_name, verify=verify, retry=False)
        return False
    except exceptions.RequestException:
        return False


def unzip_file(file_name_zip: str, file_name_extract: str) -> bool:
    path_zip = os.path.join(get_app_data_directory(), "temp", file_name_zip)
    try:
        with ZipFile(path_zip) as zip_file:
            zip_file.extract(file_name_extract, os.path.join(get_app_data_directory(), "temp"))
            return True
    except BadZipFile:
        return False


def update_list(list_id: int, sub_id: int) -> None:
    list_function_dict: dict[int, Callable[[int], list[Player] | None]] = {
        1: get_fide_list, 2: get_dsb_list, 3: get_uscf_list, 4: get_ecf_list, 5: get_oesb_list, 6: get_knsb_list,
        7: get_mcf_list, 8: get_cfr_list, 9: get_fcf_list, 10: get_nzcf_list
    }
    name = NAMES_DICT[(list_id, sub_id)]
    players_new = list_function_dict[list_id](sub_id)
    if players_new is None:
        return
    collection_uuid = players_new[0].get_uuid_associate()
    if collection_exists("", collection_uuid):
        players = DB_PLAYER.load_all("", collection_uuid)
    else:
        players = []
        DB_COLLECTION.add_list("", [Collection(name, "Players", collection_uuid)])
    DB_PLAYER.add_list("", list(set(players_new) - set(players)))
    DB_PLAYER.update_list("", list(set(players) & set(players_new)))
    DB_PLAYER.remove_list("", list(set(players) - set(players_new)))


def get_fide_list(sub_id: int) -> list[Player] | None:
    rating_list = {0: "standard", 1: "rapid", 2: "blitz"}[sub_id]
    url = f"http://ratings.fide.com/download/{rating_list}_rating_list_xml.zip"
    file_name_zip = "fide.zip"
    file_name_xml = f"{rating_list}_rating_list.xml"
    path_zip = os.path.join(get_app_data_directory(), "temp", file_name_zip)
    path_xml = os.path.join(get_app_data_directory(), "temp", file_name_xml)

    if not download_file(url, file_name_zip) or not unzip_file(file_name_zip, file_name_xml):
        return None
    data = process_xml(path_xml, ("name", "sex", "birthday", "country", "title", "rating", "fideid"))
    os.remove(path_zip)
    os.remove(path_xml)
    return [Player(
        entry[1], entry[3], entry[6], entry[2], entry[4], entry[5],
        get_uuid_from_numbers(int(entry[0]), 1, sub_id), get_uuid_from_numbers(0, 1, sub_id)
    ) for entry in data]


def get_dsb_list(sub_id: int) -> list[Player] | None:
    url = "https://dwz.svw.info/services/files/export/csv/LV-0-csv_v2.zip"
    file_name_zip = "dsb.zip"
    file_name_csv = "spieler.csv"
    path_zip = os.path.join(get_app_data_directory(), "temp", file_name_zip)
    path_csv = os.path.join(get_app_data_directory(), "temp", file_name_csv)
    certificate = os.path.join(get_app_data_directory(), "certificates", "svw-info-certificate.pem")

    if not download_file(url, file_name_zip, verify=certificate) or not unzip_file(file_name_zip, file_name_csv):
        return None
    data = process_csv(path_csv, ("Spielername", "Geschlecht", "Geburtsjahr", "FIDE-Land", "FIDE-Titel", "DWZ", "ID"))
    os.remove(path_zip)
    os.remove(path_csv)
    return [Player(
        entry[0].replace(",", ", "), entry[1], entry[2], entry[3], None if entry[4] == '-' else entry[4], entry[5],
        get_uuid_from_numbers(int(entry[6]), 2, sub_id), get_uuid_from_numbers(0, 2, sub_id)
    ) for entry in data]


def get_uscf_list(sub_id: int) -> list[Player] | None:
    rating_key = {
        0: "R_LPB_RAT", 1: "Q_LPB_RAT", 2: "B_LPB_RAT", 3: "ONL_R_RTG", 4: "ONL_Q_RTG", 5: "ONL_B_RTG"
    }[sub_id]
    url = "https://www.kingregistration.com/combineddb/db"
    file_name_zip = "uscf.zip"
    file_name_dbf = "uscffide.dbf"
    path_zip = os.path.join(get_app_data_directory(), "temp", file_name_zip)
    path_dbf = os.path.join(get_app_data_directory(), "temp", file_name_dbf)

    if not download_file(url, file_name_zip) or not unzip_file(file_name_zip, file_name_dbf):
        return None
    data: list[tuple[str, ...]] = [(
        record["MEM_NAME"], record["GENDER"], record["BYEAR"], record["FEDERATION"],
        record["TITLE"], record[rating_key], record["MEM_ID"]
    ) for record in DBF(path_dbf)]
    os.remove(path_zip)
    os.remove(path_dbf)
    return [Player(
        entry[0].replace(",", ", ").title(), entry[1], entry[2], entry[3], entry[4], entry[5],
        get_uuid_from_numbers(int(entry[6]), 3, sub_id), get_uuid_from_numbers(0, 3, sub_id)
    ) for entry in data]


def get_ecf_list(sub_id: int) -> list[Player] | None:
    rating_list = {0: "standard", 1: "rapid", 2: "blitz"}[sub_id]
    url = "https://www.ecfrating.org.uk/v2/new/api.php?v2/rating_list_csv"
    file_name = "ecf.csv"
    path = os.path.join(get_app_data_directory(), "temp", file_name)

    if not download_file(url, file_name):
        return None
    data = process_csv(path, ("full_name", "gender", "", "nation", "", f"original_{rating_list}", "ECF_code"), "utf-8")
    os.remove(path)
    return [Player(
        entry[0], entry[1], None, entry[3], None, entry[5],
        get_uuid_from_numbers(int(entry[6][:-1]), 4, sub_id), get_uuid_from_numbers(0, 4, sub_id)
    ) for entry in data]


def get_oesb_list(sub_id: int) -> list[Player] | None:
    search_text = {0: "downloads/oesb", 1: "downloads/oesbr"}[sub_id]
    url = get_link_from_url("https://www.swiss-chess.de/dt_daten.php", search_text)
    if url is None:
        return None
    url = f"https://www.swiss-chess.de/{url}"
    name = url[url.rindex('/') + 1:-4]
    file_name_zip = f"{name}.zip"
    file_name_dbf = f"{name}.LST"
    path_zip = os.path.join(get_app_data_directory(), "temp", file_name_zip)
    path_LST = os.path.join(get_app_data_directory(), "temp", file_name_dbf)

    if not download_file(url, file_name_zip, agent=False) or not unzip_file(file_name_zip, file_name_dbf):
        return None
    data_raw: list[list[str]] = [line.split(';') for line in read_file(path_LST, encoding="iso-8859-15").split('\n')]
    data: list[tuple[str, ...]] = [(
        item[0][1:-1], item[10][1:-1], item[6][1:-1], item[2][1:-1], item[5][1:-1], item[4][1:-1], item[7][1:-1]
    ) for item in data_raw[:-1]]
    os.remove(path_zip)
    os.remove(path_LST)
    return [Player(
        entry[0], entry[1], entry[2], entry[3], entry[4], entry[5],
        get_uuid_from_numbers(int(entry[6]), 5, sub_id), get_uuid_from_numbers(0, 5, sub_id)
    ) for entry in data]


def get_knsb_list(sub_id: int) -> list[Player] | None:
    url = get_link_from_url("https://schaakbond.nl/rating/downloadlijsten/", "/KNSB.zip", front='=', back='>')
    if url is None:
        return None
    file_name_zip = "KNSB.zip"
    file_name_csv = "KNSB.csv"
    path_zip = os.path.join(get_app_data_directory(), "temp", file_name_zip)
    path_csv = os.path.join(get_app_data_directory(), "temp", file_name_csv)

    if not download_file(url, file_name_zip) or not unzip_file(file_name_zip, file_name_csv):
        return None
    data = [row.split(';') for row in read_file(path_csv, encoding=None).split('\n')[1:-1]]
    os.remove(path_zip)
    os.remove(path_csv)
    return [Player(
        entry[1], entry[7].upper() or 'M', entry[6], entry[3], entry[2], entry[4],
        get_uuid_from_numbers(int(entry[0]), 6, sub_id), get_uuid_from_numbers(0, 6, sub_id)
    ) for entry in data]


def get_mcf_list(sub_id: int) -> list[Player] | None:
    url = "https://rating.malaysiachess.my/api/mcfratinglist.ashx"
    file_name = "MCFRating.xls"
    path = os.path.join(get_app_data_directory(), "temp", file_name)

    if not download_file(url, file_name):
        return None
    data = process_xls(path, ("Name", "Sex", "birthday", "FED", "Title", "rtg_nat", "ID_No"))
    os.remove(path)
    return [Player(
        entry[0].title(), entry[1], entry[2], entry[3], entry[4], int(float(entry[5])) or None,
        get_uuid_from_numbers(int(entry[6]), 7, sub_id), get_uuid_from_numbers(0, 7, sub_id)
    ) for entry in data]


def get_cfr_list(sub_id: int) -> list[Player] | None:
    rating_list = {0: "standard", 1: "rapid", 2: "blitz"}[sub_id]
    url = f"https://ratings.ruchess.ru/api/smanager_{rating_list}.csv.zip"
    file_name_zip = f"cfr_{rating_list}.zip"
    file_name_csv = f"smanager_{rating_list}.csv"
    path_zip = os.path.join(get_app_data_directory(), "temp", file_name_zip)
    path_csv = os.path.join(get_app_data_directory(), "temp", file_name_csv)

    if not download_file(url, file_name_zip) or not unzip_file(file_name_zip, file_name_csv):
        return None
    data = process_csv(path_csv, ("Name", "Sex", "Birthday", "Fed", "", "Rtg_Nat", "ID_No"), encoding="utf-8")
    data = [entry for entry in data if entry[6].isdigit()]
    os.remove(path_zip)
    os.remove(path_csv)
    return [Player(
        entry[0], entry[1] or 'M', entry[2], entry[3], entry[4], entry[5],
        get_uuid_from_numbers(int(entry[6]), 8, sub_id), get_uuid_from_numbers(0, 8, sub_id)
    ) for entry in data]


def get_fcf_list(sub_id: int) -> list[Player] | None:
    rating_list = {0: "selolista", 1: "pelolista"}[sub_id]
    url = f"http://www.shakki.net/selo/{rating_list}.xls"
    file_name = f"{rating_list}.xls"
    path = os.path.join(get_app_data_directory(), "temp", file_name)

    if not download_file(url, file_name):
        return None
    data = process_xls(path, ("Last Name", "First Name", "Sex", "Birthday", "Fed", "", "Rtg_Nat", "ID_No"))
    os.remove(path)
    return [Player(
        f"{entry[0]}, {entry[1]}", entry[2], entry[3][:4], entry[4], entry[5], int(float(entry[6])) or None,
        get_uuid_from_numbers(int(float(entry[7])), 9, sub_id), get_uuid_from_numbers(0, 9, sub_id)
    ) for entry in data]


def get_nzcf_list(sub_id: int) -> list[Player] | None:
    rating_list = {0: "Standard", 1: "Rapid"}[sub_id]
    url = get_link_from_url("https://newzealandchess.co.nz/new-zealand-ratings-list", "-Alphabetical-")
    if url is None:
        return None
    file_name = url[url.rindex('/') + 1:]
    path = os.path.join(get_app_data_directory(), "temp", file_name)

    if not download_file(url, file_name):
        return None
    data = process_xlsx(
        path, ("Name", "First_Name", "", "Year", "FIDE_Fed", "FIDE_Title", f"{rating_list}_Rating", "Code"),
        header_rows=2, header_add=["", "", "First_Name"]
    )
    os.remove(path)
    return [Player(
        f"{entry[0]}, {entry[1]}", entry[2], entry[3], entry[4], entry[5].strip(), entry[6],
        get_uuid_from_numbers(int(entry[7][:-2]), 10, sub_id), get_uuid_from_numbers(0, 10, sub_id)
    ) for entry in data]
